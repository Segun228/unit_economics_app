from math import inf
from typing import Dict, List
from rest_framework.response import Response
import pandas as pd
import zipfile
import numpy as np
import seaborn as sns
import matplotlib as mplb
mplb.use('Agg')
import logging
from pprint import pprint
from django.http.response import HttpResponseBadRequest
from .report_handlers import process_dataframe
import matplotlib.pyplot as plt
import io
from django.http import HttpResponse
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

def unit_count_bep(data):
    try:
        df = pd.DataFrame([data])
        proc = process_dataframe(df)
        if proc is None:
            raise ValueError("Error while calculating BEP")

        proc = proc.to_dict(orient="records")[0]


        k_val = proc.get("UCM")
        a_val = proc.get("FC")
        bep_val = proc.get("Required_units_to_BEP")
        

        if k_val is None or a_val is None or bep_val is None:
            raise ValueError(f"Missing required data: UCM={k_val}, FC={a_val}, Required_units_to_BEP={bep_val}")


        k = float(k_val)
        a = float(a_val)
        bep_units = int(bep_val) + 1

        x = list(range(0, 2 * bep_units))
        y1 = [-a for _ in x]
        y2 = [-a + k * xi for xi in x]

        fig, ax = plt.subplots()
        ax.plot(x, y1, label="Постоянные издержки", linewidth=2)
        ax.plot(x, y2, label="Юниты масштабирования", linewidth=2)

        ax.set_xlim(0, max(x))
        min_y = min(min(y1), min(y2))
        max_y = max(max(y1), max(y2))
        ax.set_ylim(min_y - abs(min_y)*0.2, max_y + abs(max_y)*0.2)

        ax.set_xlabel("Units")
        ax.set_ylabel("Cash flow")
        ax.set_title("Расчет точки безубыточности")
        ax.axhline(0, color='black', linewidth=0.5)
        ax.axvline(0, color='black', linewidth=0.5)
        ax.grid(False)


        ax.scatter(bep_units, 0, color='red', zorder=5, label=f"BEP = {bep_units}")
        ax.annotate(f'BEP = {bep_units}', xy=(bep_units, 0), xytext=(bep_units+1, a/4),
                    arrowprops=dict(facecolor='black', arrowstyle='->'))

        ax.legend()
        fig.subplots_adjust(left=0.15)


        buf = io.BytesIO()
        fig.savefig(buf, format='png', dpi=300, bbox_inches='tight')
        buf.seek(0)
        plt.close(fig)

        return proc, buf

    except Exception as e:
        logging.error(f"Error in unit_count_bep: {e}")
        raise

def unit_generate_report(data):
    try:
        df = pd.DataFrame([data])
        result = process_dataframe(df)
        if result is None:
            raise ValueError("Error while calculating economics")
        else:
            result_dict = result.to_dict(orient="records")[0]
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                result.to_excel(
                    writer,
                    index=False,
                    sheet_name="Текущие посты"
                )
                worksheet = writer.sheets["Текущие посты"]
                header_font = Font(bold=True)
                for col_num, column_title in enumerate(result.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                for column_cells in worksheet.columns:
                    length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                    col_letter = get_column_letter(column_cells[0].column)
                    worksheet.column_dimensions[col_letter].width = length + 4
            buffer.seek(0)
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename={result_dict.get("name", "Model")}.xlsx'
            return response

    except Exception as e:
        logging.exception(e)
        raise


def unit_count_cohort(data):
    try:
        print(data)
        result = pd.DataFrame([data])
        if result is None:
            raise ValueError("Error while calculating economics")
        else:
            result_dict = result.to_dict(orient="records")[0]
            conv = (result["customers"] / result["users"]).to_list()[0]
            result["cohort"] = 1
            result["new users"] = result["users"]
            result.loc[0, "new users"] = result.loc[0, "users"]
            result["user retention"] = 0
            result["user churn"] = 0
            result["total users"] = result.loc[0, "users"]
            result["Accumulative profit"] = 0
            result["Ballance"] = 0 - result["FC"]

            result = pd.concat([result for _ in range(24)], axis=0, ignore_index=True, copy=True)
            for i in range(1, len(result)):
                result.loc[i, "cohort"] = result.loc[i-1, "cohort"] + 1
                result.loc[i, "new users"] = result.loc[i-1, "total users"] * result.loc[i, "AGR"] 
                result.loc[i, "user retention"] = result.loc[i-1, "total users"] * result.loc[i-1, "RR"] 
                result.loc[i, "user churn"] = result.loc[i-1, "total users"] * (1 - result.loc[i-1, "RR"])
                result.loc[i, "total users"] = result.loc[i, "user retention"] + result.loc[i, "new users"]

            result["Unit"] = "User"
            result["C1"] = conv
            result["customers"] = result["total users"]*conv
            result["ARPC"] = result["AVP"] * result["APC"]
            result["ARPU"] = result["ARPC"] * result["C1"]
            result["CPA"] = result["TMS"] / result["total users"]
            result["CAC"] = result["TMS"] / result["customers"]
            result["CLTV"] = (result["AVP"] - result["COGS"]) * result["APC"] - result["COGS1s"]
            result["LTV"] = result["CLTV"] * result["C1"]
            result["ROI"] = (result["LTV"] - result["CPA"]) / result["CPA"] * 100
            result["UCM"] = result["LTV"] - result["CPA"]
            result["CCM"] = result["CLTV"] - result["CAC"]


            result["Profitable"] = np.where(result["UCM"] > 0, "YES", "NO")

            result["Revenue"] = result["ARPU"] * result["total users"]
            result["Gross_profit"] = result["CLTV"] * result["customers"]
            result["Margin"] = result["Gross_profit"] - result["TMS"]


            def calculate_required_bep(row: pd.Series):
                ucm = row.get("UCM", 0)
                if ucm > 0:
                    return row.get("FC", 0) / ucm
                return None


            result["Required_units_to_BEP"] = result.apply(calculate_required_bep, axis=1)
            
            result["BEP"] = result["Required_units_to_BEP"] * result["UCM"]
            result["Profit"] = result["Margin"] - result["FC"]

            float_cols = result.select_dtypes(include='float').columns
            result[float_cols] = result[float_cols].round(4)

            for i in range(1, len(result)):
                result.loc[i, "Accumulative profit"] = result.loc[i-1, "Accumulative profit"] + result.loc[i, "Profit"]
                result.loc[i, "Ballance"] = result.loc[i-1, "Ballance"] + result.loc[i, "Profit"] - result.loc[i, "FC"]


            result.replace([np.inf, -np.inf], np.nan, inplace=True)
            result = result.infer_objects(copy=False)
            result = result.where(pd.notnull(result), None)


            result = result.where(pd.notnull(result), None)[[
                'name', 'Unit', 'cohort',

                # 👥 Пользователи и клиенты
                'users', 'customers', 'new users', 'user retention', 'user churn', 'total users',

                # Прочее (если нужно): 'C1'
                'C1',

                # 🎯 Метрики роста и удержания
                'RR', 'AGR',

                # 💰 Финансовые метрики
                'AVP', 'APC', 'ARPC', 'ARPU', 'TMS', 'COGS', 'COGS1s', 'Gross_profit', 'Margin',

                # 📈 Экономика привлечения и монетизации
                'CPA', 'CAC', 'CLTV', 'LTV', 'ROI', 'UCM', 'CCM',

                # 🧾 Операционные расходы и прибыль
                'Revenue', 'FC', 'Profit', 'Accumulative profit', 'Ballance',

                # ⚖️ Рентабельность и BEP
                'Profitable', 'Required_units_to_BEP', 'BEP',

            ]]
            
            # Building exel

            exel_buffer = io.BytesIO()
            with pd.ExcelWriter(exel_buffer, engine='openpyxl') as writer:
                result.to_excel(
                    writer,
                    index=False,
                    sheet_name="Когортный анализ"
                )
                worksheet = writer.sheets["Когортный анализ"]
                header_font = Font(bold=True)
                for col_num, column_title in enumerate(result.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                for column_cells in worksheet.columns:
                    length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                    col_letter = get_column_letter(column_cells[0].column)
                    worksheet.column_dimensions[col_letter].width = length + 4
            exel_buffer.seek(0)
            
            # Building images
            prof_data = (result['Profit']).tolist()
            ballance_data = result['Ballance'].tolist()
            total_users_data = result['total users'].tolist()

            def build_plot(data, name, x_name, y_name):
                x = list(range(len(data)))

                fig, ax = plt.subplots()
                ax.plot(x, data, marker='.', label=name)
                ax.set_xlabel(x_name)
                ax.set_ylabel(y_name)
                ax.set_title(name)
                ax.grid(True)
                ax.axhline(0, color='black', linewidth=0.5)
                ax.axvline(0, color='black', linewidth=0.5)
                ax.legend()

                image_buf = io.BytesIO()
                fig.savefig(image_buf, format='png', dpi=300, bbox_inches='tight')
                image_buf.seek(0)
                plt.close(fig)
                return image_buf
            images = []
            images.append(build_plot(prof_data, "График маржинальной прибыли", "Период", "Прибыль"))
            images.append(build_plot(ballance_data, "График балланса", "Период", "Балланс"))
            images.append(build_plot(total_users_data, "График аудитории", "Период", "Аудитория"))


            buffer = io.BytesIO()
            with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
                zipf.writestr("report.xlsx", exel_buffer.getvalue())
                for i, img in enumerate(images, start=1):
                    zipf.writestr(f"image_{i}.png", img.getvalue())
            buffer.seek(0)
            response = HttpResponse(buffer.getvalue(), content_type='application/zip')
            response['Content-Disposition'] = 'attachment; filename="report_bundle.zip"'
            return response
    except Exception as e:
        logging.exception(e)
        raise

if __name__ == "__main__":
    dummy_data = {
        "name": "Тестовая модель",
        "users": 1000,
        "customers": 100,
        "AVP": 300,
        "APC": 2,
        "TMS": 20000,
        "COGS": 100,
        "COGS1s": 20,
        "FC": 10000,
        "RR":0.27,
        "AGR":0.12
    }
    print(unit_count_cohort(dummy_data))
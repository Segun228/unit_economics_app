from math import inf
from typing import Dict, List
from rest_framework.response import Response
import pandas as pd
import zipfile
import numpy as np
import seaborn as sns
import matplotlib as mplb
mplb.use('Agg')
import logging
from pprint import pprint
from django.http.response import HttpResponseBadRequest
from .report_handlers import process_dataframe
import matplotlib.pyplot as plt
import io
from django.http import HttpResponse
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from .report_handlers import unit_calculate_economics

def set_generate_report(data):
    calculated_units = []
    errors = []
    try:
        set_name = data.get("name", "Model Set Name")
        units = data.get("units")
        if not units or not isinstance(units, list):
            raise ValueError("Empty or invalid 'units' list")
        for i, unit in enumerate(units, start=1):
            try:
                result = unit_calculate_economics(unit)
                if result is not None:
                    result = result[0]
                    df_unit = pd.DataFrame([result])  
                    df_unit["Unit"] = unit.get("name", f"Unit_{i}") 
                    calculated_units.append(df_unit)
                else:
                    errors.append({"index": i, "error": "Calculation returned None"})
            except Exception as e:
                logging.warning(f"Ошибка при расчёте unit[{i}]: {e}")
                errors.append({"index": i, "error": str(e)})

        if not calculated_units:
            raise ValueError("No valid units for visualization")

        big_df = pd.concat(calculated_units, ignore_index=True)
        result = process_dataframe(big_df)
        if result is None:
            raise ValueError("Error while calculating economics")
        else:
            result_dict = result.to_dict(orient="records")[0]
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                result.to_excel(
                    writer,
                    index=False,
                    sheet_name="Текущие посты"
                )
                worksheet = writer.sheets["Текущие посты"]
                header_font = Font(bold=True)
                for col_num, column_title in enumerate(result.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                for column_cells in worksheet.columns:
                    length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                    col_letter = get_column_letter(column_cells[0].column)
                    worksheet.column_dimensions[col_letter].width = length + 4
            buffer.seek(0)
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename={result_dict.get("name", "Model")}.xlsx'
            return response

    except Exception as e:
        logging.exception(e)
        raise


def process_dataframe_cohort(data:pd.DataFrame):
    try:
        result = data
        if result is None:
            raise ValueError("Error while calculating economics")
        else:
            conv = (result["customers"] / result["users"]).to_list()[0]
            result["cohort"] = 1
            result["new users"] = result["users"]
            result.loc[0, "new users"] = result.loc[0, "users"]
            result["user retention"] = 0
            result["user churn"] = 0
            result["total users"] = result.loc[0, "users"]
            result["Accumulative profit"] = 0
            result["Ballance"] = 0 - result["FC"]

            result = pd.concat([result for _ in range(24)], axis=0, ignore_index=True, copy=True)
            for i in range(1, len(result)):
                result.loc[i, "cohort"] = result.loc[i-1, "cohort"] + 1
                result.loc[i, "new users"] = result.loc[i-1, "total users"] * result.loc[i, "AGR"] 
                result.loc[i, "user retention"] = result.loc[i-1, "total users"] * result.loc[i-1, "RR"] 
                result.loc[i, "user churn"] = result.loc[i-1, "total users"] * (1 - result.loc[i-1, "RR"])
                result.loc[i, "total users"] = result.loc[i, "user retention"] + result.loc[i, "new users"]

            result["Unit"] = "User"
            result["C1"] = conv
            result["customers"] = result["total users"]*conv
            result["ARPC"] = result["AVP"] * result["APC"]
            result["ARPU"] = result["ARPC"] * result["C1"]
            result["CPA"] = result["TMS"] / result["total users"]
            result["CAC"] = result["TMS"] / result["customers"]
            result["CLTV"] = (result["AVP"] - result["COGS"]) * result["APC"] - result["COGS1s"]
            result["LTV"] = result["CLTV"] * result["C1"]
            result["ROI"] = (result["LTV"] - result["CPA"]) / result["CPA"] * 100
            result["UCM"] = result["LTV"] - result["CPA"]
            result["CCM"] = result["CLTV"] - result["CAC"]


            result["Profitable"] = np.where(result["UCM"] > 0, "YES", "NO")

            result["Revenue"] = result["ARPU"] * result["total users"]
            result["Gross_profit"] = result["CLTV"] * result["customers"]
            result["Margin"] = result["Gross_profit"] - result["TMS"]


            def calculate_required_bep(row: pd.Series):
                ucm = row.get("UCM", 0)
                if ucm > 0:
                    return row.get("FC", 0) / ucm
                return None


            result["Required_units_to_BEP"] = result.apply(calculate_required_bep, axis=1)
            
            result["BEP"] = result["Required_units_to_BEP"] * result["UCM"]
            result["Profit"] = result["Margin"] - result["FC"]

            float_cols = result.select_dtypes(include='float').columns
            result[float_cols] = result[float_cols].round(4)

            for i in range(1, len(result)):
                result.loc[i, "Accumulative profit"] = result.loc[i-1, "Accumulative profit"] + result.loc[i, "Profit"]
                result.loc[i, "Ballance"] = result.loc[i-1, "Ballance"] + result.loc[i, "Profit"] - result.loc[i, "FC"]


            result.replace([np.inf, -np.inf], np.nan, inplace=True)
            result = result.infer_objects(copy=False)
            result = result.where(pd.notnull(result), None)


            result = result.where(pd.notnull(result), None)[[
                'name', 'Unit', 'cohort',
                # 👥 Пользователи и клиенты
                'users', 'customers', 'new users', 'user retention', 'user churn', 'total users',
                # Прочее (если нужно): 'C1'
                'C1',
                # 🎯 Метрики роста и удержания
                'RR', 'AGR',
                # 💰 Финансовые метрики
                'AVP', 'APC', 'ARPC', 'ARPU', 'TMS', 'COGS', 'COGS1s', 'Gross_profit', 'Margin',
                # 📈 Экономика привлечения и монетизации
                'CPA', 'CAC', 'CLTV', 'LTV', 'ROI', 'UCM', 'CCM',
                # 🧾 Операционные расходы и прибыль
                'Revenue', 'FC', 'Profit', 'Accumulative profit', 'Ballance',
                # ⚖️ Рентабельность и BEP
                'Profitable', 'Required_units_to_BEP', 'BEP',
            ]]
            return result
    except Exception as e:
        logging.exception(e)
        raise


def get_xlsx_multiple_report(data:list):
    try:
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            for ind, model in enumerate(data, start=1):
                model.to_excel(
                    writer,
                    index=False,
                    sheet_name= model.get("name", f"Model {ind}")
                )
                worksheet = writer.sheets[f"Model {ind}"]
                header_font = Font(bold=True)
                for col_num, column_title in enumerate(model.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                for column_cells in worksheet.columns:
                    length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                    col_letter = get_column_letter(column_cells[0].column)
                    worksheet.column_dimensions[col_letter].width = length + 4
        buffer.seek(0)
        return buffer
    except Exception as e:
        logging.exception(e)
        raise


def build_comparison_plots(dataframes: list[pd.DataFrame]) -> list[io.BytesIO]:
    """
    Принимает список датафреймов и строит сравнительные графики:
    1. График прибыли
    2. График баланса
    3. График аудитории

    Возвращает список изображений (в виде BytesIO).
    """
    images = []

    metrics = [
        ("Profit", "График маржинальной прибыли", "Прибыль"),
        ("Ballance", "График балланса", "Балланс"),
        ("total users", "График аудитории", "Аудитория")
    ]

    for column_name, plot_title, y_label in metrics:
        combined_df = pd.DataFrame()

        for idx, df in enumerate(dataframes):
            temp_df = pd.DataFrame({
                "Период": list(range(len(df[column_name]))),
                y_label: df[column_name].tolist(),
                "Источник": f"DF_{idx+1}"  # Название серии
            })
            combined_df = pd.concat([combined_df, temp_df], ignore_index=True)

        fig, ax = plt.subplots(figsize=(10, 6))
        sns.lineplot(
            data=combined_df,
            x="Период",
            y=y_label,
            hue="Источник",
            marker="o",
            ax=ax
        )

        ax.set_title(plot_title)
        ax.set_xlabel("Период")
        ax.set_ylabel(y_label)
        ax.grid(True)
        ax.axhline(0, color='black', linewidth=0.5)

        image_buf = io.BytesIO()
        fig.savefig(image_buf, format='png', dpi=300, bbox_inches='tight')
        image_buf.seek(0)
        plt.close(fig)
        images.append(image_buf)

    return images




def set_count_cohort(data):
    calculated_units = []
    errors = []
    try:
        set_name = data.get("name", "Model Set Name")
        units = data.get("units")
        if not units or not isinstance(units, list):
            raise ValueError("Empty or invalid 'units' list")
        for i, unit in enumerate(units, start=1):
            try:
                result = unit_calculate_economics(unit)
                if result is not None:
                    result = result[0]
                    df_unit = pd.DataFrame([result])  
                    df_unit["Unit"] = unit.get("name", f"Unit_{i}") 
                    calculated_units.append(df_unit)
                else:
                    errors.append({"index": i, "error": "Calculation returned None"})
            except Exception as e:
                logging.warning(f"Ошибка при расчёте unit[{i}]: {e}")
                errors.append({"index": i, "error": str(e)})
        if not calculated_units:
            raise ValueError("No valid units for visualization")
        calculated_units = list(map(process_dataframe_cohort, calculated_units))

        buf_xlsx = get_xlsx_multiple_report(calculated_units)
        images = build_comparison_plots(calculated_units)

        buffer = io.BytesIO()
        with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
            zipf.writestr("report.xlsx", buf_xlsx.getvalue())
            for i, img in enumerate(images, start=1):
                zipf.writestr(f"image_{i}.png", img.getvalue())
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename="report_bundle.zip"'
        return response
    except Exception as e:
        logging.exception(e)
        raise
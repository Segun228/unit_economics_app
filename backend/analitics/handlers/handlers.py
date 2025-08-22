import pandas as pd
from api.models import ModelSet, UnitModel
import logging
from django.forms.models import model_to_dict
from io import BytesIO
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework.exceptions import bad_request
from django.http import HttpResponseBadRequest


def add_posts_file(data, request=None, name="New set via XLSX", description=""):
    df = pd.read_excel(data)
    units = df.to_dict(orient="records")
    errors = []

    user = getattr(request, 'user')

    new_set = ModelSet.objects.create(
        name=name,
        description=description,
        user=user
    )

    for i, unit in enumerate(units, start=1):
        try:
            UnitModel.objects.create(
                name=unit.get("unit name", ""),
                users=unit.get("users", ""),
                model_set=new_set,
                customers=unit.get("customers", ""),
                AVP=unit.get("AVP", ""),
                APC=unit.get("APC", ""),
                TMS=unit.get("TMS", ""),
                COGS=unit.get("COGS", ""),
                COGS1s=unit.get("COGS1s", ""),
                FC=unit.get("FC", ""),
                user=user
            )
        except Exception as e:
            logging.error(f"Ошибка в строке {i}: {e}")
            errors.append({"row": i, "error": str(e)})

    return {
        "success": True,
        "errors": errors,
        "created": len(units) - len(errors),
        "skipped": len(errors)
    }



def get_xlsx_report(units, sets):
    if not units or not sets:
        return HttpResponseBadRequest("Нет данных для отчета")
    units_df = pd.DataFrame(units)[[
        "model_set_id",
        "name",
        "users",
        "customers",
        "AVP",
        "APC",
        "TMS",
        "COGS",
        "COGS1s",
        "FC"
    ]]
    sets_df = pd.DataFrame(sets)[["id", "name", "description"]]
    result = pd.merge(
        left=units_df,
        right=sets_df,
        left_on="model_set_id",
        right_on="id",
        how="inner",
        suffixes=("_unit", "_set")
    )[["name_set", "description",	"name_unit", "users", "customers", "AVP", "APC", "TMS", "COGS", "COGS1s", "FC"]]
    result.columns = ["set name", "description",	"unit name", "users", "customers", "AVP", "APC", "TMS", "COGS", "COGS1s", "FC"]
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        result.to_excel(
            writer,
            index=False,
            sheet_name="Текущие посты"
        )
        worksheet = writer.sheets["Текущие посты"]
        header_font = Font(bold=True)
        for col_num, column_title in enumerate(result.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            col_letter = get_column_letter(column_cells[0].column)
            worksheet.column_dimensions[col_letter].width = length + 4
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=products.xlsx'
    return response

